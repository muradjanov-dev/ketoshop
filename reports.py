"""
Excel report generator — orders + per-line-item breakdown.

Used by the admin "📊 Excel hisobot" entry. Returns an in-memory .xlsx
file built with openpyxl so we don't touch the filesystem.
"""
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from database import get_orders_for_export, get_all_cost_prices, format_local_dt
from locales import get_delivery_method_name, get_order_status


# ----- Styling -----------------------------------------------------

_HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)
_TOTAL_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
_TOTAL_FONT = Font(bold=True)


def _write_header(ws, headers: list[str]) -> None:
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"


def _autosize(ws, min_widths: list[int] | None = None) -> None:
    """Auto-size columns based on content; cap at 50 chars."""
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            v = cell.value
            if v is None:
                continue
            length = len(str(v))
            if length > max_len:
                max_len = length
        width = min(max_len + 2, 50)
        if min_widths and col_idx <= len(min_widths):
            width = max(width, min_widths[col_idx - 1])
        ws.column_dimensions[col_letter].width = width


# ----- Builder -----------------------------------------------------

def _fmt_dt(dt) -> str:
    """Format an order timestamp in Asia/Tashkent local time for the Excel
    report. Stored values are naive UTC; format_local_dt applies the +5h
    shift consistently with the rest of the bot UI."""
    return format_local_dt(dt, "%Y-%m-%d %H:%M")


def _payment_label(method: str | None, lang: str) -> str:
    if method == "cash":
        return "💵 Naqd" if lang == "uz" else "💵 Наличные"
    if method == "online":
        return "💳 Onlayn" if lang == "uz" else "💳 Онлайн"
    return "—"


def _source_label(source: str | None, lang: str) -> str:
    if source == "manual":
        return "📝 Qo'lda" if lang == "uz" else "📝 Вручную"
    if source == "webapp":
        return "📱 Mini App"
    return "🤖 Bot"


PERIOD_LABELS = {
    "today": {"uz": "Bugun",          "ru": "Сегодня"},
    "7d":    {"uz": "Oxirgi 7 kun",   "ru": "Последние 7 дней"},
    "30d":   {"uz": "Oxirgi 30 kun",  "ru": "Последние 30 дней"},
    "all":   {"uz": "Barcha vaqt",    "ru": "За всё время"},
}


def _orders_sheet(ws, orders: list[dict], lang: str, cost_map: dict[int, float]) -> None:
    ws.title = "Buyurtmalar" if lang == "uz" else "Заказы"
    headers_uz = [
        "Buyurtma №", "Sana", "Mijoz", "Telefon",
        "Mahsulot", "Miqdor", "Birlik narx", "Qator summa",
        "Asl narx", "Foyda",
        "Buyurtma jami", "To'lov", "Yetkazib berish", "Holat", "Manba",
        "Manzil", "Izoh", "Tasdiqlangan", "Yo'lda", "Yetkazilgan",
    ]
    headers_ru = [
        "Заказ №", "Дата", "Клиент", "Телефон",
        "Товар", "Кол-во", "Цена за ед.", "Сумма строки",
        "Себестоимость", "Прибыль",
        "Сумма заказа", "Оплата", "Доставка", "Статус", "Источник",
        "Адрес", "Комментарий", "Подтверждён", "В пути", "Доставлен",
    ]
    headers = headers_uz if lang == "uz" else headers_ru
    _write_header(ws, headers)

    row = 2
    for o in orders:
        items = o.get("items_data") or []
        if not items:
            items = [{"name": "—", "quantity": 1, "price": float(o.get("total") or 0)}]
        # One row per item — accountants want each line separately.
        for it in items:
            qty = float(it.get("quantity") or 0)
            price = float(it.get("price") or 0)
            line_total = qty * price
            pid = it.get("product_id")
            unit_cost = cost_map.get(int(pid), 0.0) if pid else 0.0
            line_cost = unit_cost * qty
            line_profit = line_total - line_cost

            ws.cell(row=row, column=1,  value=int(o["id"]))
            ws.cell(row=row, column=2,  value=_fmt_dt(o.get("created_at")))
            ws.cell(row=row, column=3,  value=o.get("customer_name") or "—")
            ws.cell(row=row, column=4,  value=o.get("phone") or "—")
            ws.cell(row=row, column=5,  value=it.get("name") or "—")
            ws.cell(row=row, column=6,  value=qty)
            ws.cell(row=row, column=7,  value=int(price))
            ws.cell(row=row, column=8,  value=int(line_total))
            ws.cell(row=row, column=9,  value=int(unit_cost))
            ws.cell(row=row, column=10, value=int(line_profit))
            ws.cell(row=row, column=11, value=int(float(o.get("total") or 0)))
            ws.cell(row=row, column=12, value=_payment_label(o.get("payment_method"), lang))
            ws.cell(row=row, column=13, value=get_delivery_method_name(o.get("delivery_method"), lang))
            ws.cell(row=row, column=14, value=get_order_status(o.get("status") or "pending", lang))
            ws.cell(row=row, column=15, value=_source_label(o.get("source"), lang))
            ws.cell(row=row, column=16, value=o.get("address") or "—")
            ws.cell(row=row, column=17, value=o.get("address_note") or "")
            ws.cell(row=row, column=18, value=_fmt_dt(o.get("confirmed_at")))
            ws.cell(row=row, column=19, value=_fmt_dt(o.get("shipped_at")))
            ws.cell(row=row, column=20, value=_fmt_dt(o.get("delivered_at")))

            # Right-align numeric columns
            for col in (1, 6, 7, 8, 9, 10, 11):
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="right")
            for col in range(1, 21):
                ws.cell(row=row, column=col).border = _BORDER
            row += 1
    _autosize(ws, min_widths=[10, 17, 18, 14, 24, 8, 12, 12, 12, 12, 14, 14, 18, 16, 12, 28, 22, 17, 17, 17])


def _summary_sheet(ws, orders: list[dict], lang: str, period: str,
                   cost_map: dict[int, float]) -> None:
    ws.title = "Xulosa" if lang == "uz" else "Сводка"

    delivered = [o for o in orders if o.get("status") == "delivered"]
    revenue = sum(float(o.get("total") or 0) for o in delivered)
    aov = (revenue / len(delivered)) if delivered else 0

    # Cost & profit on delivered orders only — accounting truth is "what we
    # actually shipped", not pending/cancelled. Lines without product_id
    # (legacy or imported) contribute 0 cost.
    total_cost = 0.0
    for o in delivered:
        for it in (o.get("items_data") or []):
            pid = it.get("product_id")
            if not pid:
                continue
            qty = float(it.get("quantity") or 0)
            total_cost += cost_map.get(int(pid), 0.0) * qty
    profit = revenue - total_cost
    margin_pct = (profit / revenue * 100) if revenue else 0

    by_status: dict[str, int] = {}
    for o in orders:
        s = o.get("status") or "pending"
        by_status[s] = by_status.get(s, 0) + 1

    by_source: dict[str, int] = {}
    for o in orders:
        s = (o.get("source") or "bot")
        by_source[s] = by_source.get(s, 0) + 1

    period_label = PERIOD_LABELS.get(period, {}).get(lang, period)

    rows: list[tuple[str, object]] = [
        (("Davr"             if lang == "uz" else "Период"), period_label),
        (("Jami buyurtmalar" if lang == "uz" else "Всего заказов"), len(orders)),
        (("Yetkazilgan"      if lang == "uz" else "Доставлено"), len(delivered)),
        (("Daromad (so'm)"   if lang == "uz" else "Выручка (сум)"), int(revenue)),
        (("Asl narx jami"    if lang == "uz" else "Себестоимость"),  int(total_cost)),
        (("Foyda (so'm)"     if lang == "uz" else "Прибыль (сум)"),  int(profit)),
        (("Foyda %"          if lang == "uz" else "Маржа %"),        round(margin_pct, 1)),
        (("O'rtacha chek"    if lang == "uz" else "Средний чек"), int(aov)),
        ("", ""),
        ((" Holat bo'yicha"  if lang == "uz" else " По статусу"), ""),
    ]
    for status, count in sorted(by_status.items()):
        rows.append((get_order_status(status, lang), count))
    rows.append(("", ""))
    rows.append(((" Manba bo'yicha" if lang == "uz" else " По источнику"), ""))
    for source, count in sorted(by_source.items()):
        rows.append((_source_label(source, lang), count))

    # Top block (period + revenue/cost/profit/aov) is highlighted
    HIGHLIGHT_TOP_ROWS = 8
    for i, (k, v) in enumerate(rows, start=1):
        a = ws.cell(row=i, column=1, value=k)
        b = ws.cell(row=i, column=2, value=v)
        if k and not v and ("bo'yicha" in str(k) or "По" in str(k)):
            a.font = _TOTAL_FONT
            a.fill = _TOTAL_FILL
        if i <= HIGHLIGHT_TOP_ROWS:
            a.font = _TOTAL_FONT
            b.font = _TOTAL_FONT
            a.fill = _TOTAL_FILL
            b.fill = _TOTAL_FILL
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22


# ----- Public API --------------------------------------------------

async def generate_orders_excel(period: str = "all", lang: str = "uz") -> tuple[BytesIO, str]:
    """Build an .xlsx report for the given period. Returns (buffer, filename)."""
    orders = await get_orders_for_export(period)
    # One DB call → in-memory map; avoids N+1 lookups while iterating items.
    cost_map = await get_all_cost_prices()

    wb = Workbook()
    _orders_sheet(wb.active, orders, lang, cost_map)
    summary = wb.create_sheet()
    _summary_sheet(summary, orders, lang, period, cost_map)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    today = datetime.now().strftime("%Y-%m-%d")
    filename = f"safran_orders_{period}_{today}.xlsx"
    return buf, filename
